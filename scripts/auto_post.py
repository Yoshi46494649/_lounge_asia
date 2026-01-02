#!/usr/bin/env python3
"""Automated Lounge Asia blog post generator and publisher."""

from __future__ import annotations

import argparse
import dataclasses
import json
import logging
import math
import os
import random
import re
import shutil
import subprocess
import sys
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Sequence

os.environ.setdefault("PYTHONUTF8", "1")

import markdown
import yaml
from openai import OpenAI
from openpyxl import load_workbook
from slugify import slugify
from tenacity import retry, stop_after_attempt, wait_exponential

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    from backports.zoneinfo import ZoneInfo  # type: ignore


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG_PATH = ROOT / "scripts" / "config.yaml"
SYSTEM_PROMPT = """\
You are Lounge Asia's in-house blog copywriter, blending SEO expertise with brand storytelling.
Follow every rule below and respond strictly with a single JSON object in UTF-8:

Brand Tone & Persona:
- Tone: Sophisticated, calm, reassuring. Highlight the theme 窶廚omfort and Romance窶・
- Audience: Adults (25-40) with Asian cultural interests or backgrounds, living in Australia.
- Language: Simple English (middle school reading level). Use Australian English spellings.

Mandatory Information (include somewhere in the article):
- Host: Lounge Asia (Brisbane窶冱 largest Asian community with 500+ members).
- Theme: Comfort and Romance.
- Tagline: Connecting hearts across Asian cultures.
- Venue: The Stock Exchange Hotel 窶・Level 2.
- Dress code: Smart Casual.
- Explain why exchanging contact details during the event is paused: to keep the event fair and comfortable.

Privacy & Compliance:
- Never include personal data (names, phone numbers, emails, votes, etc.).
- Do not invent testimonials or identifiable stories.

Article Structure (Markdown inside `content`):
- Provide an H1 title (handled separately in `title`).
- Include a short table of contents.
- Use H2 and H3 headings to organise sections.
- Include a dedicated CTA section encouraging bookings or joining the waitlist.
- Add an 窶廾fficial Information窶・or 窶廛isclaimer窶・section covering event policies.
- Add a 窶彝elated Articles窶・section heading (content will be injected later).

Output JSON keys (all are required):
- title: SEO title that starts with the provided keyword.
- meta_description: ~150 characters meta description optimised for clicks.
- category: Must match the supplied category.
- content: Full article body in Markdown format.
"""

REVIEW_PROMPT = """\
You are reviewing a Lounge Asia blog article for compliance.
Rules:
- Maintain Australian English spelling and simple language.
- Ensure mandatory details (host, theme, tagline, venue, dress code, contact-exchange policy) appear naturally.
- Keep tone elegant and comforting.
- Respect privacy (no personal data).
- Keep headings logical. Table of contents, CTA, Disclaimer/Official Information, Related Articles section headings must exist.

Return a JSON object with:
- "is_compliant": boolean (true only if every rule is satisfied).
- "content": If changes are required, rewrite the full article Markdown. Otherwise return the original unchanged content.
- "notes": Short array of strings summarising adjustments (optional but useful).
"""


class AutoPostError(RuntimeError):
    """Custom error type for auto post failures."""


def setup_logging(verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="[%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


def load_yaml_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise AutoPostError(f"Config file not found: {path}")
    with path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    return data


def read_guidelines(path: Path) -> str:
    if not path.exists():
        raise AutoPostError(f"Guideline file not found: {path}")
    return path.read_text(encoding="utf-8")


@dataclass
class KeywordCandidate:
    row_index: int
    keyword: str
    category: str
    difficulty: Any
    extra: dict[str, Any]


def _normalize_for_sort(value: Any, ascending: bool) -> Any:
    if value is None:
        return math.inf if ascending else -math.inf
    if isinstance(value, (int, float)):
        return value if ascending else -value
    text = str(value).strip()
    if text == "":
        return math.inf if ascending else -math.inf
    try:
        # Remove commas or percentage signs
        cleaned = text.replace(",", "").replace("%", "")
        number = float(cleaned)
        return number if ascending else -number
    except ValueError:
        return text.lower() if ascending else f"~{text.lower()}"


def load_keyword_candidates(config: dict[str, Any]) -> list[KeywordCandidate]:
    ks_conf = config.get("keyword_source", {})
    path = ks_conf.get("path")
    sheet_name = ks_conf.get("sheet_name")
    keyword_col = ks_conf.get("keyword_column")
    category_col = ks_conf.get("category_column")
    difficulty_col = ks_conf.get("difficulty_column")
    posted_col = ks_conf.get("posted_flag_column")
    posted_truthy = {str(v).strip().lower() for v in ks_conf.get("posted_truthy_values", [])}
    priority_columns = ks_conf.get("priority_columns", [])

    if not path:
        raise AutoPostError("`keyword_source.path` is not configured")

    workbook_path = ROOT / path
    if not workbook_path.exists():
        raise AutoPostError(f"Keyword workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    ws = None
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {str(name): idx for idx, name in enumerate(headers)}

    required_columns = [keyword_col, category_col, posted_col]
    missing = [col for col in required_columns if col not in header_map]
    if missing:
        raise AutoPostError(f"Missing expected columns in keyword sheet: {', '.join(missing)}")

    candidates: list[KeywordCandidate] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {str(headers[i]): row[i] for i in range(len(headers))}
        posted_value = row_dict.get(posted_col)
        if posted_value is not None and str(posted_value).strip().lower() in posted_truthy:
            continue
        keyword = str(row_dict.get(keyword_col, "")).strip()
        category = str(row_dict.get(category_col, "")).strip()
        if not keyword or not category:
            continue
        difficulty_value = row_dict.get(difficulty_col) if difficulty_col else None
        extra = {k: v for k, v in row_dict.items() if k not in {keyword_col, category_col}}
        candidates.append(
            KeywordCandidate(
                row_index=row_idx,
                keyword=keyword,
                category=category,
                difficulty=difficulty_value,
                extra=extra,
            )
        )

    if not candidates:
        raise AutoPostError("No available keywords remaining in workbook.")

    def sort_key(candidate: KeywordCandidate):
        keys: list[Any] = []
        for entry in priority_columns:
            if isinstance(entry, dict):
                name = entry.get("name")
                ascending = bool(entry.get("ascending", True))
            else:
                name = entry
                ascending = True
            value = candidate.extra.get(name)
            keys.append(_normalize_for_sort(value, ascending))
        return tuple(keys)

    if priority_columns:
        candidates.sort(key=sort_key)

    return candidates


def choose_image_path(category: str, content_conf: dict[str, Any]) -> tuple[Path, str]:
    image_root = ROOT / content_conf.get("image_root", "")
    if not image_root.exists():
        raise AutoPostError(f"Image root directory not found: {image_root}")

    mapping = content_conf.get("image_category_map", {})
    folder_name = mapping.get(category) or mapping.get(category.title()) or mapping.get("default")
    if not folder_name:
        raise AutoPostError(f"No image folder mapping for category '{category}'")

    folder_path = image_root / folder_name
    if not folder_path.exists():
        raise AutoPostError(f"Image folder not found: {folder_path}")

    images = [p for p in folder_path.iterdir() if p.is_file() and p.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}]
    if not images:
        raise AutoPostError(f"No images available in {folder_path}")

    selected = random.choice(images)
    web_base = content_conf.get("image_web_base", "/blog-images")
    rel_parts = [web_base.rstrip("/"), folder_name.strip("/"), selected.name]
    web_path = "/".join(part.strip("/") for part in rel_parts)
    return selected, web_path


def insert_related_section(markdown_body: str, links: list[dict[str, str]]) -> str:
    lines = ["## Related Articles", ""]
    for link in links:
        lines.append(f"- [{link['title']}]({link['url']})")
    new_section = "\n".join(lines).strip()

    pattern = re.compile(r"##\s+(Related Articles|\u95a2\u9023\u8a18\u4e8b)\b[\s\S]*?(?=\n##\s|\n#\s|$)", re.IGNORECASE)
    if pattern.search(markdown_body):
        return pattern.sub(new_section, markdown_body).strip() + "\n"
    return markdown_body.rstrip() + "\n\n" + new_section + "\n"


def build_related_links(category: str, config_links: list[dict[str, Any]], minimum: int = 3) -> list[dict[str, str]]:
    preferred = [link for link in config_links if category in link.get("categories", [])]
    remaining = [link for link in config_links if link not in preferred]

    chosen: list[dict[str, str]] = []
    for source in (preferred, remaining):
        for link in source:
            if len(chosen) >= minimum:
                break
            chosen.append({"title": link["title"], "url": link["url"]})
        if len(chosen) >= minimum:
            break
    return chosen


def compute_read_time(markdown_body: str, min_minutes: int, max_minutes: int) -> int:
    words = re.findall(r"\w+", markdown_body)
    estimated = max(min_minutes, math.ceil(len(words) / 200))
    return min(estimated, max_minutes)


def prepare_front_matter(
    article: dict[str, Any],
    keyword: KeywordCandidate,
    slug: str,
    hero_web_path: str,
    content_conf: dict[str, Any],
) -> dict[str, Any]:
    now = datetime.now(ZoneInfo(content_conf.get("timezone", "Australia/Brisbane")))
    read_time = compute_read_time(article["content"], content_conf.get("min_read_minutes", 4), content_conf.get("max_read_minutes", 12))
    excerpt = article.get("meta_description") or article["content"].split("\n", 1)[0].strip()
    front_matter = {
        "title": article["title"].strip(),
        "slug": slug,
        "keyword": keyword.keyword,
        "category": article["category"].strip(),
        "meta_description": article["meta_description"].strip(),
        "excerpt": excerpt,
        "published_at": now.strftime("%Y-%m-%d"),
        "read_time_minutes": read_time,
        "hero_image": hero_web_path,
        "hero_image_alt": content_conf.get("hero_alt_template", "{title}").format(title=article["title"].strip()),
        "author": content_conf.get("default_author", "Lounge Asia Team"),
        "tags": content_conf.get("default_tags", []),
        "cta": "Reserve your place with Lounge Asia today.",
        "featured": False,
    }
    return front_matter


def render_markdown(front_matter: dict[str, Any], hero_web_path: str, content_body: str) -> str:
    fm_yaml = yaml.safe_dump(front_matter, sort_keys=False, allow_unicode=True).strip()
    hero_alt = front_matter.get("hero_image_alt", front_matter["title"])
    hero_block = f"![{hero_alt}]({hero_web_path})"
    body = content_body.strip()
    return f"---\n{fm_yaml}\n---\n\n{hero_block}\n\n{body}\n"


def write_markdown_file(path: Path, content: str, dry_run: bool) -> None:
    if dry_run:
        logging.info("[dry-run] Skipping write to %s", path)
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    logging.info("Saved article to %s", path)


def sanitize_js_string(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')

def find_block_bounds(source: str, start_index: int, open_char: str, close_char: str) -> tuple[int, int]:
    level = 0
    block_start = None
    for offset, ch in enumerate(source[start_index:], start_index):
        if ch == open_char:
            if level == 0:
                block_start = offset
            level += 1
        elif ch == close_char:
            level -= 1
            if level == 0:
                if block_start is None:
                    break
                return block_start, offset + 1
    raise AutoPostError(f"Failed to locate matching {close_char} in bundle.")


def append_js_array_entry(source: str, var_name: str, entry: str, slug: str) -> str:
    anchor = f"const {var_name}="
    start = source.find(anchor)
    if start == -1:
        raise AutoPostError(f"Could not find array '{var_name}' in legacy bundle.")
    array_start = source.find("[", start)
    if array_start == -1:
        raise AutoPostError(f"Array start not found for '{var_name}'.")
    block_start, block_end = find_block_bounds(source, array_start, "[", "]")
    block = source[block_start:block_end]
    if f'slug:"{slug}"' in block:
        raise AutoPostError(f"Legacy bundle already contains slug '{slug}'.")
    new_block = block[:-1] + "," + entry + "]"
    return source[:block_start] + new_block + source[block_end:]


def append_js_object_entry(source: str, var_name: str, entry: str, slug: str) -> str:
    anchor = f"const {var_name}="
    start = source.find(anchor)
    if start == -1:
        raise AutoPostError(f"Could not find object '{var_name}' in legacy bundle.")
    obj_start = source.find("{", start)
    if obj_start == -1:
        raise AutoPostError(f"Object start not found for '{var_name}'.")
    block_start, block_end = find_block_bounds(source, obj_start, "{", "}")
    block = source[block_start:block_end]
    if f'"{slug}":' in block:
        raise AutoPostError(f"Legacy bundle already contains slug '{slug}'.")
    new_block = block[:-1] + entry + "}"
    return source[:block_start] + new_block + source[block_end:]


def ensure_category_present(source: str, category: str) -> str:
    anchor = '$y=["All"'
    start = source.find(anchor)
    if start == -1:
        logging.warning("Category list anchor not found in legacy bundle.")
        return source
    array_start = source.find("[", start)
    block_start, block_end = find_block_bounds(source, array_start, "[", "]")
    block = source[block_start:block_end]
    if f'"{category}"' in block:
        return source
    new_block = block[:-1] + f',"{sanitize_js_string(category)}"]'
    return source[:block_start] + new_block + source[block_end:]


def prepare_legacy_markdown(markdown_body: str) -> str:
    lines = markdown_body.splitlines()
    filtered: list[str] = []
    skip_contents = False
    skip_section = False
    for line in lines:
        stripped = line.strip()
        if not filtered and (stripped.startswith("![") or stripped.startswith("# ")):
            continue
        if stripped == "## Table of Contents":
            skip_contents = True
            continue
        if skip_contents:
            if stripped.startswith("## "):
                skip_contents = False
            else:
                continue
        if stripped == "## Related Articles":
            skip_section = True
            continue
        if skip_section and stripped.startswith("## "):
            skip_section = False
        if skip_section:
            continue
        filtered.append(line)
    return "\n".join(filtered).strip()


def markdown_to_html_body(markdown_text: str) -> str:
    if not markdown_text:
        return ""
    return markdown.markdown(markdown_text, extensions=["extra"])


def format_html_for_bundle(html: str) -> str:
    if not html:
        return ""
    stripped = html.strip()
    indented = stripped.replace("\r\n", "\n").replace("\n", "\n      ")
    return f"\n      {indented}\n    "


def extract_excerpt_from_html(html: str, fallback: str, max_length: int = 220) -> str:
    if not html:
        return fallback
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return fallback
    if len(text) <= max_length:
        return text
    trimmed = text[:max_length].rstrip()
    cut = trimmed.rfind(" ")
    if cut > 50:
        trimmed = trimmed[:cut]
    return trimmed + "..."


def update_legacy_site(front_matter: dict[str, Any], markdown_body: str, config: dict[str, Any]) -> None:
    if not config.get("enabled"):
        return

    bundle_path = ROOT / config["bundle_path"]
    if not bundle_path.exists():
        raise AutoPostError(f"Legacy bundle not found: {bundle_path}")

    filtered_markdown = prepare_legacy_markdown(markdown_body)
    html_body = markdown_to_html_body(filtered_markdown)
    formatted_html = format_html_for_bundle(html_body)
    excerpt = extract_excerpt_from_html(html_body, front_matter.get("meta_description", ""))

    slug = front_matter["slug"]
    title = sanitize_js_string(front_matter["title"])
    category = front_matter["category"]
    image_map = config.get("image_map", {})
    image_path = image_map.get(category, image_map.get("default", "blog/speed-dating/speed-dating-hero.jpg"))
    image_literal = sanitize_js_string(image_path)
    read_time_label = f"{front_matter['read_time_minutes']} min read"
    read_time_literal = sanitize_js_string(read_time_label)
    excerpt_literal = sanitize_js_string(excerpt)
    author_literal = sanitize_js_string(front_matter.get("author", "Lounge Asia Team"))

    published_at = front_matter["published_at"]
    try:
        display_date = datetime.strptime(published_at, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        display_date = published_at

    bundle_text = bundle_path.read_text(encoding="utf-8")

    array_anchor = "const vs="
    array_start = bundle_text.find(array_anchor)
    if array_start == -1:
        raise AutoPostError("Could not locate post listing array in legacy bundle.")
    array_bracket_start = bundle_text.find("[", array_start)
    arr_start, arr_end = find_block_bounds(bundle_text, array_bracket_start, "[", "]")
    array_block = bundle_text[arr_start:arr_end]
    existing_ids = [int(m.group(1)) for m in re.finditer(r"id:(\d+)", array_block)]
    next_id = (max(existing_ids) + 1) if existing_ids else 1
    featured_slugs = set(config.get("featured_slugs", []))
    featured_literal = "!0" if slug in featured_slugs else "!1"

    array_entry = (
        f'{{id:{next_id},slug:"{slug}",title:"{title}",excerpt:"{excerpt_literal}",'
        f'category:"{sanitize_js_string(category)}",date:"{display_date}",readTime:"{read_time_literal}",'
        f'image:"{image_literal}",featured:{featured_literal}}}'
    )

    object_entry = (
        f',"{slug}":{{title:"{title}",excerpt:"{excerpt_literal}",category:"{sanitize_js_string(category)}",'
        f'date:"{display_date}",readTime:"{read_time_literal}",image:"{image_literal}",author:"{author_literal}",'
        f"content:`{formatted_html}`}}"
    )

    updated_text = append_js_array_entry(bundle_text, "vs", array_entry, slug)
    updated_text = ensure_category_present(updated_text, category)
    updated_text = append_js_object_entry(updated_text, "Gi", object_entry, slug)

    bundle_path.write_text(updated_text, encoding="utf-8")
    logging.info("Updated legacy bundle with slug '%s'", slug)

    template_source = config.get("template_source")
    output_dir_value = config.get("output_dir")
    if template_source and output_dir_value:
        template_path = ROOT / template_source
        if not template_path.exists():
            raise AutoPostError(f"Legacy template not found: {template_path}")
        slug_dir = (ROOT / output_dir_value) / slug
        slug_dir.mkdir(parents=True, exist_ok=True)
        target_file = slug_dir / "index.html"
        if not target_file.exists():
            shutil.copyfile(template_path, target_file)
            logging.info("Created legacy HTML shell at %s", target_file)
    else:
        logging.warning("Legacy template configuration incomplete; skipped HTML shell creation.")


def update_keyword_sheet(candidate: KeywordCandidate, config: dict[str, Any], dry_run: bool) -> None:
    if dry_run:
        logging.info("[dry-run] Skipping keyword sheet update")
        return

    ks_conf = config["keyword_source"]
    workbook_path = ROOT / ks_conf["path"]
    wb = load_workbook(workbook_path)
    sheet_name = ks_conf.get("sheet_name")
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    posted_col = ks_conf["posted_flag_column"]
    posted_idx = None
    for idx, header in enumerate(headers, start=1):
        if header == posted_col:
            posted_idx = idx
            break
    if posted_idx is None:
        raise AutoPostError(f"Posted flag column '{posted_col}' not found when updating sheet.")

    template = ks_conf.get("posted_marker_template", "Posted {date}")
    date_format = ks_conf.get("posted_date_format", "%Y-%m-%d")
    now = datetime.now().strftime(date_format)
    marker = template.format(date=now)
    ws.cell(row=candidate.row_index, column=posted_idx, value=marker)
    wb.save(workbook_path)
    logging.info("Updated keyword sheet for row %s", candidate.row_index)


def sanitise_output(text: str) -> str:
    return text.strip().encode("ascii", "backslashreplace").decode("ascii")


def run_subprocess(command: Sequence[str], cwd: Path | None = None, env: dict[str, str] | None = None) -> None:
    logging.debug("Running command: %s (cwd=%s)", " ".join(command), cwd or ".")
    cmd = list(command)
    if os.name == "nt":
        executable = cmd[0]
        if executable in {"npm", "npx", "semgrep"} and not executable.endswith(".cmd"):
            resolved = shutil.which(f"{executable}.cmd") or shutil.which(executable) or executable
            cmd[0] = resolved
    run_env = {**os.environ, **(env or {})}
    run_env.setdefault("PYTHONUTF8", "1")
    run_env.setdefault("PYTHONIOENCODING", "utf-8")
    result = subprocess.run(
        cmd,
        cwd=cwd,
        env=run_env,
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if result.stdout:
        logging.debug(sanitise_output(result.stdout))
    if result.returncode != 0:
        if result.stderr:
            logging.error(sanitise_output(result.stderr))
        raise AutoPostError(f"Command {' '.join(command)} failed with exit code {result.returncode}")


def run_ci_pipeline(ci_commands: list[dict[str, Any]], dry_run: bool) -> None:
    if dry_run:
        logging.info("[dry-run] Skipping CI pipeline")
        return
    for entry in ci_commands:
        name = entry.get("name", "step")
        cmd = entry.get("command")
        cwd = entry.get("cwd")
        if not cmd:
            logging.warning("CI step '%s' has no command; skipping", name)
            continue
        cwd_path = ROOT / cwd if cwd and cwd != "." else ROOT
        logging.info("Running CI step: %s", name)
        try:
            run_subprocess([str(part) for part in cmd], cwd=cwd_path)
        except AutoPostError as exc:
            raise AutoPostError(f"CI step '{name}' failed: {exc}") from exc


def stage_and_commit(files: Iterable[Path], commit_message: str, branch: str, dry_run: bool) -> None:
    if dry_run:
        logging.info("[dry-run] Skipping git commit")
        return

    paths = [str(path.relative_to(ROOT)) for path in files]
    run_subprocess(["git", "add", *paths], cwd=ROOT)
    status = subprocess.run(["git", "status", "--porcelain"], cwd=ROOT, capture_output=True, text=True, check=False)
    if not status.stdout.strip():
        logging.info("No git changes detected; skipping commit.")
        return

    run_subprocess(["git", "commit", "-m", commit_message], cwd=ROOT)
    run_subprocess(["git", "push", "origin", branch], cwd=ROOT)
    logging.info("Changes pushed to %s", branch)


def ensure_openai_client(timeout: int) -> OpenAI:
    if "OPENAI_API_KEY" not in os.environ:
        raise AutoPostError("OPENAI_API_KEY environment variable is not set.")
    return OpenAI(timeout=timeout)


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=2, min=2, max=20))
def call_openai(client: OpenAI, model: str, system_prompt: str, user_prompt: str, temperature: float, max_tokens: int) -> str:
    response = client.responses.create(
        model=model,
        input=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=temperature,
        max_output_tokens=max_tokens,
    )
    if not response.output_text:
        raise AutoPostError("OpenAI response did not include text output.")
    return response.output_text.strip()


def generate_article(client: OpenAI, config: dict[str, Any], keyword: KeywordCandidate, guidelines: str) -> dict[str, Any]:
    openai_conf = config.get("openai", {})
    model = openai_conf["model_draft"]
    tone_guidance = config.get("content", {}).get("tone_guidance", [])
    tone_instructions = ""
    if tone_guidance:
        guidance_lines = "\n".join(f"- {item}" for item in tone_guidance if isinstance(item, str) and item.strip())
        if guidance_lines:
            tone_instructions = "\n\nTone & engagement adjustments:\n" + guidance_lines + "\n"
    prompt = (
        f"{guidelines}\n\n"
        "Generate a Lounge Asia blog article using the above rules.\n"
        f"繧ｿ繝ｼ繧ｲ繝・ヨ繧ｭ繝ｼ繝ｯ繝ｼ繝・ {keyword.keyword}\n"
        f"繧ｫ繝・ざ繝ｪ: {keyword.category}\n"
        f"{tone_instructions}"
    )
    logging.info("Generating article for keyword '%s'", keyword.keyword)
    output = call_openai(
        client=client,
        model=model,
        system_prompt=SYSTEM_PROMPT,
        user_prompt=prompt,
        temperature=openai_conf.get("temperature", 0.4),
        max_tokens=openai_conf.get("max_output_tokens", 2400),
    )
    try:
        return json.loads(output)
    except json.JSONDecodeError:
        cleaned = output.strip()
        if cleaned.startswith("```"):
            segments = cleaned.split("```")
            if len(segments) >= 3:
                payload = segments[1]
                if payload.lstrip().startswith("json"):
                    payload = payload.split("\n", 1)[1]
                try:
                    return json.loads(payload)
                except json.JSONDecodeError:
                    pass
        logging.error("Failed to parse OpenAI response:\n%s", output)
        raise AutoPostError("Draft generation did not return valid JSON.")


def review_article(client: OpenAI, config: dict[str, Any], article: dict[str, Any], guidelines: str) -> dict[str, Any]:
    openai_conf = config.get("openai", {})
    model = openai_conf["model_review"]
    tone_guidance = config.get("content", {}).get("tone_guidance", [])
    tone_instructions = ""
    if tone_guidance:
        guidance_lines = "\n".join(f"- {item}" for item in tone_guidance if isinstance(item, str) and item.strip())
        if guidance_lines:
            tone_instructions = "\n\nTone & engagement adjustments:\n" + guidance_lines + "\n"
    prompt = (
        f"{guidelines}\n\n"
        "Review the following article JSON and ensure it complies with every rule.\n"
        "If adjustments are needed, rewrite the `content` section directly.\n"
        f"{tone_instructions}"
        "```\n"
        f"{json.dumps(article, ensure_ascii=False, indent=2)}\n"
        "```"
    )
    logging.info("Running compliance review")
    output = call_openai(
        client=client,
        model=model,
        system_prompt=REVIEW_PROMPT,
        user_prompt=prompt,
        temperature=0.2,
        max_tokens=openai_conf.get("max_output_tokens", 2400),
    )
    try:
        review_payload = json.loads(output)
    except json.JSONDecodeError:
        cleaned = output.strip()
        if cleaned.startswith("```"):
            segments = cleaned.split("```")
            if len(segments) >= 3:
                payload = segments[1]
                if payload.lstrip().startswith("json"):
                    payload = payload.split("\n", 1)[1]
                try:
                    review_payload = json.loads(payload)
                except json.JSONDecodeError:
                    logging.error("Review response parsing failed:\n%s", output)
                    raise AutoPostError("Review stage returned invalid JSON.")
                else:
                    pass
        else:
            logging.error("Review response parsing failed:\n%s", output)
            raise AutoPostError("Review stage returned invalid JSON.")
        if 'review_payload' not in locals():
            logging.error("Review response parsing failed:\n%s", output)
            raise AutoPostError("Review stage returned invalid JSON.")

    if not isinstance(review_payload, dict) or "content" not in review_payload:
        raise AutoPostError("Review payload missing required keys: content/is_compliant")

    content = review_payload["content"]
    if not isinstance(content, str) or not content.strip():
        raise AutoPostError("Review stage produced empty content.")
    article["content"] = content
    if "notes" in review_payload:
        logging.info("Review notes: %s", review_payload["notes"])
    return article


def build_commit_message(template: str, article: dict[str, Any], slug: str) -> str:
    return template.format(title=article["title"], slug=slug)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Automate Lounge Asia blog post creation.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH, help="Path to YAML config.")
    parser.add_argument("--keyword", type=str, default=None, help="Keyword to force (must exist in workbook).")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Generate outputs without writing files, updating Excel, or committing.",
    )
    parser.add_argument("--skip-ci", action="store_true", help="Skip CI command execution.")
    parser.add_argument("--skip-git", action="store_true", help="Skip git add/commit/push even if enabled.")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
    parser.add_argument("--mock-json", type=Path, default=None, help="Use pre-generated article JSON instead of calling OpenAI.")
    args = parser.parse_args(argv)

    setup_logging(verbose=args.verbose)
    config = load_yaml_config(args.config)
    guidelines = read_guidelines(ROOT / config["content"]["guideline_path"])
    candidates = load_keyword_candidates(config)

    if args.keyword:
        keyword = next((c for c in candidates if c.keyword.lower() == args.keyword.lower()), None)
        if not keyword:
            raise AutoPostError(f"Keyword '{args.keyword}' not found or already posted.")
    else:
        keyword = candidates[0]

    mock_article_path = args.mock_json
    mock_article: dict[str, Any] | None = None
    if mock_article_path:
        if not mock_article_path.exists():
            raise AutoPostError(f"Mock JSON not found: {mock_article_path}")
        with mock_article_path.open("r", encoding="utf-8") as fh:
            mock_article = json.load(fh)
        logging.info("Using mock article JSON from %s", mock_article_path)

    client: OpenAI | None = None
    if mock_article is None:
        client = ensure_openai_client(config["openai"].get("timeout_seconds", 180))
        article = generate_article(client, config, keyword, guidelines)
        article = review_article(client, config, article, guidelines)
    else:
        article = mock_article

    required_fields = {
        "title": str,
        "meta_description": str,
        "category": str,
        "content": str,
    }
    for field, field_type in required_fields.items():
        value = article.get(field)
        if not isinstance(value, field_type) or not str(value).strip():
            raise AutoPostError(f"Article is missing required field '{field}'.")

    if article.get("category") != keyword.category:
        logging.warning("Model returned category '%s' but expected '%s'. Overriding.", article.get("category"), keyword.category)
        article["category"] = keyword.category

    related_links = build_related_links(article["category"], config.get("internal_links", []))
    article["content"] = insert_related_section(article["content"], related_links)

    content_conf = config["content"]
    _, hero_web_path = choose_image_path(article["category"], content_conf)
    slug_base = slugify(article["title"], lowercase=True)
    slug = slug_base or slugify(keyword.keyword, lowercase=True)
    output_dir = ROOT / content_conf["output_dir"]
    output_path = output_dir / f"{slug}.md"
    if output_path.exists():
        timestamp = datetime.now().strftime("%Y%m%d%H%M")
        slug = f"{slug}-{timestamp}"
        output_path = output_dir / f"{slug}.md"
        logging.warning("Slug already existed. Using %s", slug)

    front_matter = prepare_front_matter(article, keyword, slug, hero_web_path, content_conf)
    markdown_output = render_markdown(front_matter, hero_web_path, article["content"])

    write_markdown_file(output_path, markdown_output, args.dry_run)
    if not args.dry_run:
        legacy_conf = config.get("legacy_site", {})
        if legacy_conf.get("enabled"):
            update_legacy_site(front_matter, article["content"], legacy_conf)
    update_keyword_sheet(keyword, config, args.dry_run)

    if not args.skip_ci:
        run_ci_pipeline(config.get("ci_commands", []), args.dry_run)

    git_conf = config.get("git", {})
    if git_conf.get("enable", True) and not args.skip_git:
        commit_message = build_commit_message(git_conf.get("commit_message_template", "feat: add blog post {title}"), article, slug)
        stage_and_commit(
            files=[output_path, ROOT / config["keyword_source"]["path"]],
            commit_message=commit_message,
            branch=git_conf.get("branch", "main"),
            dry_run=args.dry_run,
        )

    logging.info("Automation complete. Keyword '%s' processed.", keyword.keyword)
    return 0


if __name__ == "__main__":  # pragma: no cover
    try:
        raise SystemExit(main())
    except AutoPostError as exc:
        logging.error("Automation failed: %s", exc)
        raise SystemExit(1)








